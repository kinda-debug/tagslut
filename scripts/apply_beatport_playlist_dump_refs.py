#!/usr/bin/env python3
"""Apply duration refs from a Beatport playlist dump + share pages.

Input dump format expected:
- A text file containing HTTP response capture where the playlist list JSON
  appears after a marker line like: `Response: 200 ()`.

Behavior:
1) Parse playlist IDs from dump.
2) Fetch `https://www.beatport.com/playlists/share/<id>` and parse __NEXT_DATA__.
3) Extract first page track payloads and build refs:
   - beatport track id -> length_ms
   - isrc -> median(length_ms) across observed rows
4) Insert refs with `INSERT OR IGNORE`.
5) Recompute duration status for rows from one XLSX playlist.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from collections import defaultdict
from datetime import datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

import httpx
from openpyxl import load_workbook

NEXT_DATA_RE = re.compile(
    r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>',
    re.IGNORECASE | re.DOTALL,
)
ISRC_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}\d{7}$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Apply Beatport playlist dump refs and recompute playlist statuses.")
    parser.add_argument("--dump", type=Path, required=True, help="Beatport playlists dump path.")
    parser.add_argument("--db", type=Path, required=True, help="SQLite DB path.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Target playlist XLSX (Path column).")
    parser.add_argument(
        "--source-label",
        default="beatport_playlists_dump",
        help="ref_source value for inserted refs.",
    )
    parser.add_argument("--timeout", type=float, default=25.0, help="HTTP timeout seconds.")
    return parser.parse_args()


def _extract_playlist_ids_from_dump(path: Path) -> list[str]:
    text = path.read_text(encoding="utf-8", errors="replace")
    marker = "Response: 200 ()"
    start = text.find(marker)
    if start == -1:
        raise RuntimeError("Could not find 'Response: 200 ()' marker in dump.")
    payload = text[start + len(marker) :]
    brace = payload.find("{")
    if brace == -1:
        raise RuntimeError("Could not find JSON object start after marker.")
    decoder = json.JSONDecoder()
    obj, _ = decoder.raw_decode(payload[brace:])
    results = obj.get("results", []) if isinstance(obj, dict) else []
    out: list[str] = []
    for item in results:
        if not isinstance(item, dict):
            continue
        pid = item.get("id")
        if pid is None:
            continue
        out.append(str(pid))
    return out


def _extract_playlist_tracks(client: httpx.Client, playlist_id: str) -> list[dict]:
    url = f"https://www.beatport.com/playlists/share/{playlist_id}"
    response = client.get(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0 Safari/537.36"
            )
        },
    )
    if response.status_code != 200:
        return []
    match = NEXT_DATA_RE.search(response.text)
    if not match:
        return []
    try:
        next_data = json.loads(match.group(1))
    except Exception:
        return []

    queries = (
        next_data.get("props", {})
        .get("pageProps", {})
        .get("dehydratedState", {})
        .get("queries", [])
    )
    if not isinstance(queries, list):
        return []

    for query in queries:
        key = query.get("queryKey")
        if not isinstance(key, list):
            continue
        if not any(f"catalog-playlist-{playlist_id}-page=" in str(x) for x in key):
            continue
        data = query.get("state", {}).get("data", {})
        results = data.get("results", []) if isinstance(data, dict) else []
        if isinstance(results, list):
            return results
    return []


def _load_xlsx_paths(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    sheet = wb["Tracks"] if "Tracks" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [c for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "Path" not in header:
        raise RuntimeError(f"'Path' column not found in {xlsx}")
    idx = header.index("Path")
    out = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        val = row[idx]
        if val:
            out.append(str(val))
    return out


def _safe_json(payload: str | None) -> dict:
    if not payload:
        return {}
    try:
        parsed = json.loads(payload)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _extract_tag_value(meta: dict, keys: list[str]) -> str | None:
    lowered = {str(k).lower(): v for k, v in meta.items()}
    for key in keys:
        raw = lowered.get(key.lower())
        if raw is None:
            continue
        if isinstance(raw, list):
            if not raw:
                continue
            value = str(raw[0]).strip()
        else:
            value = str(raw).strip()
        if value:
            return value
    return None


def _lookup_ref(conn: sqlite3.Connection, beatport_id: str | None, isrc: str | None) -> tuple[int | None, str | None, str | None]:
    if beatport_id:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (beatport_id,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], beatport_id
    if isrc:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (isrc,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], isrc
    return None, None, None


def main() -> int:
    args = parse_args()
    dump_path = args.dump.expanduser().resolve()
    db_path = args.db.expanduser().resolve()
    xlsx_path = args.xlsx.expanduser().resolve()
    if not dump_path.exists():
        raise SystemExit(f"Dump not found: {dump_path}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")
    if not xlsx_path.exists():
        raise SystemExit(f"XLSX not found: {xlsx_path}")

    playlist_ids = _extract_playlist_ids_from_dump(dump_path)
    print(f"Playlist IDs from dump: {len(playlist_ids)}")

    client = httpx.Client(timeout=args.timeout, follow_redirects=True)

    beatport_refs: dict[str, int] = {}
    isrc_durations: dict[str, list[int]] = defaultdict(list)

    fetched = 0
    with_tracks = 0
    for idx, pid in enumerate(playlist_ids, start=1):
        rows = _extract_playlist_tracks(client, pid)
        fetched += 1
        if rows:
            with_tracks += 1
        for item in rows:
            if not isinstance(item, dict):
                continue
            track = item.get("track")
            if not isinstance(track, dict):
                continue
            track_id = track.get("id")
            length_ms = track.get("length_ms")
            isrc = str(track.get("isrc") or "").strip().upper()
            if track_id is not None and isinstance(length_ms, int) and length_ms > 0:
                beatport_refs[str(track_id)] = int(length_ms)
            if ISRC_RE.match(isrc) and isinstance(length_ms, int) and length_ms > 0:
                isrc_durations[isrc].append(int(length_ms))
        if idx % 10 == 0 or idx == len(playlist_ids):
            print(f"[fetch {idx}/{len(playlist_ids)}] share_pages_with_tracks={with_tracks}")

    client.close()

    isrc_refs = {}
    for isrc, values in isrc_durations.items():
        vals = sorted(values)
        isrc_refs[isrc] = vals[len(vals) // 2]

    print(f"Beatport track-id refs extracted: {len(beatport_refs)}")
    print(f"ISRC refs extracted:             {len(isrc_refs)}")

    now_iso = datetime.now(timezone.utc).isoformat()
    ok_max_ms = 2000
    warn_max_ms = 8000
    duration_version = "duration_v1_ok2_warn8"

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    try:
        inserted = 0
        for ref_id, duration_ms in beatport_refs.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO track_duration_refs
                    (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
                VALUES (?, 'beatport', ?, ?, ?)
                """,
                (ref_id, duration_ms, args.source_label, now_iso),
            )
            inserted += conn.total_changes > 0

        for ref_id, duration_ms in isrc_refs.items():
            conn.execute(
                """
                INSERT OR IGNORE INTO track_duration_refs
                    (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
                VALUES (?, 'isrc', ?, ?, ?)
                """,
                (ref_id, duration_ms, args.source_label, now_iso),
            )

        print(f"track_duration_refs insert attempts complete.")

        # Recompute statuses for target XLSX paths.
        paths = _load_xlsx_paths(xlsx_path)
        placeholders = ",".join(["?"] * len(paths))
        rows = conn.execute(
            f"""
            SELECT path, metadata_json, duration_measured_ms
            FROM files
            WHERE path IN ({placeholders})
            """,
            paths,
        ).fetchall()
        by_path = {r["path"]: r for r in rows}

        recomputed = 0
        for idx, path in enumerate(paths, start=1):
            row = by_path.get(path)
            if row is None:
                continue
            meta = _safe_json(row["metadata_json"])
            beatport_id = _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            isrc = _extract_tag_value(meta, ["ISRC", "TSRC", "isrc", "tsrc"])
            if isrc:
                isrc = isrc.upper()
            measured = int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None
            ref_ms, ref_source, ref_track_id = _lookup_ref(conn, beatport_id, isrc)
            delta = measured - ref_ms if (measured is not None and ref_ms is not None) else None
            if delta is None:
                status = "unknown"
            else:
                ad = abs(delta)
                if ad <= ok_max_ms:
                    status = "ok"
                elif ad <= warn_max_ms:
                    status = "warn"
                else:
                    status = "fail"

            conn.execute(
                """
                UPDATE files
                SET duration_ref_ms = ?,
                    duration_ref_source = ?,
                    duration_ref_track_id = ?,
                    duration_ref_updated_at = ?,
                    duration_delta_ms = ?,
                    duration_status = ?,
                    duration_check_version = ?
                WHERE path = ?
                """,
                (ref_ms, ref_source, ref_track_id, now_iso, delta, status, duration_version, path),
            )
            recomputed += 1
            if idx % 200 == 0 or idx == len(paths):
                print(f"[recompute {idx}/{len(paths)}] updated={recomputed}")

        conn.commit()
        print("Done.")
        print(f"  recomputed_rows: {recomputed}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())

