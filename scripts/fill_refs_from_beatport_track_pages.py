#!/usr/bin/env python3
"""Fill duration refs from Beatport public track pages for one XLSX playlist."""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import httpx
from openpyxl import load_workbook

NEXT_DATA_RE = re.compile(
    r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>',
    re.IGNORECASE | re.DOTALL,
)
ISRC_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}\d{7}$")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Fill beatport refs from public track pages.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Playlist XLSX with Path column.")
    parser.add_argument("--db", type=Path, required=True, help="SQLite DB path.")
    parser.add_argument(
        "--source-label",
        default="beatport_track_page",
        help="ref_source value for inserted refs.",
    )
    parser.add_argument("--timeout", type=float, default=25.0, help="HTTP timeout seconds.")
    return parser.parse_args()


def _safe_json(payload: str | None) -> dict:
    if not payload:
        return {}
    try:
        parsed = json.loads(payload)
        return parsed if isinstance(parsed, dict) else {}
    except Exception:
        return {}


def _extract_tag_value(meta: dict, keys: list[str]) -> str | None:
    lowered = {str(k).lower(): v for k, v in meta.items()}
    for key in keys:
        raw = lowered.get(key.lower())
        if raw is None:
            continue
        if isinstance(raw, list):
            if not raw:
                continue
            text = str(raw[0]).strip()
        else:
            text = str(raw).strip()
        if text:
            return text
    return None


def _load_paths(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["Tracks"] if "Tracks" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "Path" not in header:
        raise RuntimeError(f"'Path' column not found in {xlsx}")
    idx = header.index("Path")
    return [str(row[idx]) for row in ws.iter_rows(min_row=2, values_only=True) if row and row[idx]]


def _beatport_track_payload(client: httpx.Client, track_id: str) -> dict | None:
    url = f"https://www.beatport.com/track/-/{track_id}"
    response = client.get(
        url,
        headers={
            "User-Agent": (
                "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0 Safari/537.36"
            )
        },
    )
    if response.status_code != 200:
        return None
    match = NEXT_DATA_RE.search(response.text)
    if not match:
        return None
    try:
        payload = json.loads(match.group(1))
    except Exception:
        return None
    queries = (
        payload.get("props", {})
        .get("pageProps", {})
        .get("dehydratedState", {})
        .get("queries", [])
    )
    if not isinstance(queries, list):
        return None
    target = f"track-{track_id}"
    for query in queries:
        key = query.get("queryKey")
        if isinstance(key, list) and any(str(item) == target for item in key):
            data = query.get("state", {}).get("data")
            if isinstance(data, dict):
                return data
    return None


def _lookup_ref(conn: sqlite3.Connection, beatport_id: str | None, isrc: str | None) -> tuple[int | None, str | None, str | None]:
    if beatport_id:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (beatport_id,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], beatport_id
    if isrc:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (isrc,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], isrc
    return None, None, None


def main() -> int:
    args = parse_args()
    xlsx = args.xlsx.expanduser().resolve()
    db_path = args.db.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"XLSX not found: {xlsx}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    paths = _load_paths(xlsx)
    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    now_iso = datetime.now(timezone.utc).isoformat()
    ok_max_ms = 2000
    warn_max_ms = 8000
    duration_version = "duration_v1_ok2_warn8"

    try:
        placeholders = ",".join(["?"] * len(paths))
        rows = conn.execute(
            f"""
            SELECT path, duration_status, duration_measured_ms, metadata_json, beatport_id
            FROM files
            WHERE path IN ({placeholders})
            """,
            paths,
        ).fetchall()
        by_path = {r["path"]: r for r in rows}

        unknown = [p for p in paths if p in by_path and by_path[p]["duration_status"] == "unknown"]
        print(f"Unknown rows in scope: {len(unknown)}")

        # Collect unresolved beatport IDs.
        target_ids: list[str] = []
        for path in unknown:
            row = by_path[path]
            meta = _safe_json(row["metadata_json"])
            beatport_id = _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            if not beatport_id and row["beatport_id"]:
                beatport_id = str(row["beatport_id"]).strip()
            if not beatport_id or not beatport_id.isdigit():
                continue
            has_ref = conn.execute(
                "SELECT 1 FROM track_duration_refs WHERE ref_id = ?",
                (beatport_id,),
            ).fetchone()
            if has_ref:
                continue
            target_ids.append(beatport_id)

        target_ids = sorted(set(target_ids))
        print(f"Unresolved beatport IDs to fetch: {len(target_ids)}")

        client = httpx.Client(timeout=args.timeout, follow_redirects=True)
        inserted = 0
        fetch_ok = 0
        fetch_fail = 0
        for idx, beatport_id in enumerate(target_ids, start=1):
            data = _beatport_track_payload(client, beatport_id)
            if not data:
                fetch_fail += 1
                continue
            length_ms = data.get("length_ms")
            if not isinstance(length_ms, int) or length_ms <= 0:
                fetch_fail += 1
                continue

            conn.execute(
                """
                INSERT OR IGNORE INTO track_duration_refs
                    (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
                VALUES (?, 'beatport', ?, ?, ?)
                """,
                (beatport_id, int(length_ms), args.source_label, now_iso),
            )
            inserted += conn.total_changes > 0
            fetch_ok += 1

            isrc = str(data.get("isrc") or "").strip().upper()
            if ISRC_RE.match(isrc):
                conn.execute(
                    """
                    INSERT OR IGNORE INTO track_duration_refs
                        (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
                    VALUES (?, 'isrc', ?, ?, ?)
                    """,
                    (isrc, int(length_ms), args.source_label, now_iso),
                )

            if idx % 50 == 0 or idx == len(target_ids):
                print(f"[fetch {idx}/{len(target_ids)}] ok={fetch_ok} fail={fetch_fail}")

        client.close()

        # Recompute statuses for all scope rows present in DB.
        recomputed = 0
        for idx, path in enumerate(paths, start=1):
            row = by_path.get(path)
            if row is None:
                continue
            meta = _safe_json(row["metadata_json"])
            beatport_id = _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            if not beatport_id and row["beatport_id"]:
                beatport_id = str(row["beatport_id"]).strip()
            isrc = _extract_tag_value(meta, ["ISRC", "TSRC", "isrc", "tsrc"])
            if isrc:
                isrc = isrc.strip().upper()
            measured = int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None
            ref_ms, ref_source, ref_track_id = _lookup_ref(conn, beatport_id, isrc)
            delta = measured - ref_ms if (measured is not None and ref_ms is not None) else None
            if delta is None:
                status = "unknown"
            else:
                ad = abs(delta)
                if ad <= ok_max_ms:
                    status = "ok"
                elif ad <= warn_max_ms:
                    status = "warn"
                else:
                    status = "fail"

            conn.execute(
                """
                UPDATE files
                SET duration_ref_ms = ?,
                    duration_ref_source = ?,
                    duration_ref_track_id = ?,
                    duration_ref_updated_at = ?,
                    duration_delta_ms = ?,
                    duration_status = ?,
                    duration_check_version = ?
                WHERE path = ?
                """,
                (ref_ms, ref_source, ref_track_id, now_iso, delta, status, duration_version, path),
            )
            recomputed += 1
            if idx % 200 == 0 or idx == len(paths):
                print(f"[recompute {idx}/{len(paths)}] updated={recomputed}")

        conn.commit()
        print("Done.")
        print(f"  fetched_ok:      {fetch_ok}")
        print(f"  fetched_fail:    {fetch_fail}")
        print(f"  recomputed_rows: {recomputed}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())

