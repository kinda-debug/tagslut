#!/usr/bin/env python3
"""Targeted unknown-duration reassessment for one playlist workbook.

Workflow:
1) Load playlist paths from XLSX ("Path" column in "Tracks" sheet).
2) For unknown rows in DB, resolve duration refs from trusted IDs only:
   - beatport_id (tag/db column) via Beatport fetch-by-id
   - ISRC (tag/canonical) via provider ISRC search
3) Upsert refs into track_duration_refs.
4) Recompute duration status for playlist rows using DB metadata + refs.
"""

from __future__ import annotations

import argparse
import json
import re
import sqlite3
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook

from tagslut.metadata.auth import TokenManager
from tagslut.metadata.providers.beatport import BeatportProvider
from tagslut.metadata.providers.spotify import SpotifyProvider
from tagslut.metadata.providers.tidal import TidalProvider


ISRC_SPLIT_RE = re.compile(r"[;,/\\]|\s+")
ISRC_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}\d{7}$")
VARIANT_WORDS = (
    "extended",
    "ext",
    "remix",
    "radio edit",
    "club mix",
    "dub",
    "edit",
    "version",
    "instrumental",
    "rework",
)


@dataclass
class Candidate:
    duration_ms: int
    title: str
    provider: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Reassess unknown duration statuses for one XLSX playlist.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Playlist XLSX path.")
    parser.add_argument("--db", type=Path, required=True, help="SQLite DB path.")
    parser.add_argument(
        "--providers",
        default="beatport,spotify,tidal",
        help="ISRC provider order (default: beatport,spotify,tidal).",
    )
    parser.add_argument(
        "--max-spread-ms",
        type=int,
        default=10000,
        help="Max allowed spread among provider durations before skip (default: 10000).",
    )
    parser.add_argument(
        "--variant-guard-delta-ms",
        type=int,
        default=90000,
        help="Block likely variant mismatch if local title/path implies version but provider does not.",
    )
    parser.add_argument(
        "--source-label",
        default="playlist_unknown_provider_guarded",
        help="ref_source label for inserted refs.",
    )
    return parser.parse_args()


def _safe_json(payload: str | None) -> dict:
    if not payload:
        return {}
    try:
        parsed = json.loads(payload)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _normalize_isrc_tokens(value) -> list[str]:
    if value is None:
        return []
    values = value if isinstance(value, list) else [value]
    out: list[str] = []
    for item in values:
        for token in ISRC_SPLIT_RE.split(str(item).strip().upper()):
            token = token.strip()
            if ISRC_RE.match(token):
                out.append(token)
    seen = set()
    uniq = []
    for token in out:
        if token in seen:
            continue
        seen.add(token)
        uniq.append(token)
    return uniq


def _extract_tag_value(meta: dict, keys: list[str]) -> str | None:
    lowered = {str(k).lower(): v for k, v in meta.items()}
    for key in keys:
        raw = lowered.get(key.lower())
        if raw is None:
            continue
        if isinstance(raw, list):
            if not raw:
                continue
            text = str(raw[0]).strip()
        else:
            text = str(raw).strip()
        if text:
            return text
    return None


def _has_variant_marker(text: str) -> bool:
    lowered = (text or "").lower()
    return any(token in lowered for token in VARIANT_WORDS)


def _choose_duration(candidates: list[Candidate], max_spread_ms: int) -> tuple[int | None, str]:
    if not candidates:
        return None, "no_candidates"
    values = sorted(c.duration_ms for c in candidates)
    spread = values[-1] - values[0]
    if spread > max_spread_ms:
        return None, f"ambiguous_spread_{spread}"
    return values[len(values) // 2], ""


def _variant_guard(
    measured_ms: int | None,
    local_text: str,
    provider_titles: Iterable[str],
    provider_duration_ms: int,
    min_delta_ms: int,
) -> tuple[bool, str]:
    if measured_ms is None:
        return False, ""
    provider_has_variant = any(_has_variant_marker(t) for t in provider_titles)
    if provider_has_variant:
        return False, ""
    if _has_variant_marker(local_text) and abs(measured_ms - provider_duration_ms) >= min_delta_ms:
        return True, "variant_mismatch_guard"
    return False, ""


def _load_playlist_paths(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    sheet = wb["Tracks"] if "Tracks" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [c for c in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "Path" not in header:
        raise SystemExit(f"'Path' column not found in {xlsx}")
    idx = header.index("Path")
    out: list[str] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = row[idx]
        if value:
            out.append(str(value))
    return out


def _query_spotify_isrc(provider: SpotifyProvider, isrc: str) -> list[Candidate]:
    out: list[Candidate] = []
    for track in provider.search_by_isrc(isrc):
        if track.isrc and track.isrc.upper() != isrc.upper():
            continue
        if track.duration_ms is None:
            continue
        out.append(Candidate(int(track.duration_ms), str(track.title or ""), "spotify"))
    return out


def _query_tidal_isrc(provider: TidalProvider, isrc: str) -> list[Candidate]:
    out: list[Candidate] = []
    for track in provider.search(isrc, limit=20):
        if not track.isrc or track.isrc.upper() != isrc.upper():
            continue
        if track.duration_ms is None:
            continue
        out.append(Candidate(int(track.duration_ms), str(track.title or ""), "tidal"))
    return out


def _query_beatport_isrc(provider: BeatportProvider, isrc: str) -> list[Candidate]:
    out: list[Candidate] = []
    for track in provider.search_by_isrc(isrc):
        if track.isrc and track.isrc.upper() != isrc.upper():
            continue
        if track.duration_ms is None:
            continue
        out.append(Candidate(int(track.duration_ms), str(track.title or ""), "beatport"))
    return out


def _query_beatport_id(provider: BeatportProvider, beatport_id: str) -> list[Candidate]:
    if not beatport_id.isdigit():
        return []
    track = provider.fetch_by_id(beatport_id)
    if track is None or track.duration_ms is None:
        return []
    return [Candidate(int(track.duration_ms), str(track.title or ""), "beatport")]


def _lookup_ref(conn: sqlite3.Connection, beatport_id: str | None, isrc_tokens: list[str]) -> tuple[int | None, str | None, str | None]:
    if beatport_id:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (beatport_id,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], beatport_id
    for isrc in isrc_tokens:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (isrc,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], isrc
    return None, None, None


def main() -> int:
    args = parse_args()
    xlsx = args.xlsx.expanduser().resolve()
    db_path = args.db.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"XLSX not found: {xlsx}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    providers_order = [p.strip().lower() for p in args.providers.split(",") if p.strip()]
    if not providers_order:
        raise SystemExit("No providers configured.")
    enabled = set(providers_order)

    paths = _load_playlist_paths(xlsx)
    print(f"Playlist paths: {len(paths)}")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row

    tm = TokenManager()
    beatport = BeatportProvider(tm)
    spotify = SpotifyProvider(tm)
    tidal = TidalProvider(tm)

    now_iso = datetime.now(timezone.utc).isoformat()
    ok_max_ms = 2000
    warn_max_ms = 8000
    duration_version = "duration_v1_ok2_warn8"

    try:
        placeholders = ",".join(["?"] * len(paths))
        rows = conn.execute(
            f"""
            SELECT
                path, metadata_json, duration_status, duration_measured_ms,
                beatport_id, canonical_isrc
            FROM files
            WHERE path IN ({placeholders})
            """,
            paths,
        ).fetchall()
        by_path = {r["path"]: r for r in rows}

        unknown_paths = [p for p in paths if p in by_path and (by_path[p]["duration_status"] == "unknown")]
        print(f"Unknown rows in DB for playlist: {len(unknown_paths)}")

        inserted_refs = 0
        skipped_no_ids = 0
        skipped_ambiguous = 0
        skipped_guard = 0
        beatport_id_cache: dict[str, list[Candidate]] = {}
        isrc_cache: dict[tuple[str, str], list[Candidate]] = {}

        for idx, path in enumerate(unknown_paths, start=1):
            row = by_path[path]
            meta = _safe_json(row["metadata_json"])
            measured_ms = int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None

            title = _extract_tag_value(meta, ["title", "TITLE"]) or ""
            local_text = f"{title} {path}"

            beatport_id = _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            if not beatport_id and row["beatport_id"]:
                beatport_id = str(row["beatport_id"]).strip()

            isrc_tokens = _normalize_isrc_tokens(
                meta.get("isrc") or meta.get("ISRC") or meta.get("tsrc") or meta.get("TSRC")
            )
            if row["canonical_isrc"]:
                isrc_tokens = _normalize_isrc_tokens(isrc_tokens + [str(row["canonical_isrc"])])

            if not beatport_id and not isrc_tokens:
                skipped_no_ids += 1
                continue

            # Skip provider lookup if ref already present for any ID.
            ref_ms, _, _ = _lookup_ref(conn, beatport_id, isrc_tokens)
            if ref_ms is not None:
                continue

            candidates: list[Candidate] = []
            titles: list[str] = []
            ref_id = None
            ref_type = None

            # Prefer beatport_id when available (only if beatport provider enabled).
            if beatport_id and "beatport" in enabled:
                if beatport_id not in beatport_id_cache:
                    beatport_id_cache[beatport_id] = _query_beatport_id(beatport, beatport_id)
                cands = beatport_id_cache[beatport_id]
                if cands:
                    candidates.extend(cands)
                    titles.extend(c.title for c in cands if c.title)
                    ref_id = beatport_id
                    ref_type = "beatport"

            if not candidates and isrc_tokens:
                target_isrc = isrc_tokens[0]
                for pname in providers_order:
                    cache_key = (pname, target_isrc)
                    if cache_key in isrc_cache:
                        cands = isrc_cache[cache_key]
                    elif pname == "beatport":
                        cands = _query_beatport_isrc(beatport, target_isrc)
                        isrc_cache[cache_key] = cands
                    elif pname == "spotify":
                        cands = _query_spotify_isrc(spotify, target_isrc)
                        isrc_cache[cache_key] = cands
                    elif pname == "tidal":
                        cands = _query_tidal_isrc(tidal, target_isrc)
                        isrc_cache[cache_key] = cands
                    else:
                        continue
                    candidates.extend(cands)
                    titles.extend(c.title for c in cands if c.title)
                    if cands:
                        # stop at first provider with results (priority order)
                        break
                if candidates:
                    ref_id = target_isrc
                    ref_type = "isrc"

            if not candidates or not ref_id or not ref_type:
                skipped_ambiguous += 1
                continue

            chosen_ms, note = _choose_duration(candidates, int(args.max_spread_ms))
            if chosen_ms is None:
                skipped_ambiguous += 1
                continue

            blocked, _ = _variant_guard(
                measured_ms=measured_ms,
                local_text=local_text,
                provider_titles=titles,
                provider_duration_ms=chosen_ms,
                min_delta_ms=int(args.variant_guard_delta_ms),
            )
            if blocked:
                skipped_guard += 1
                continue

            conn.execute(
                """
                INSERT OR REPLACE INTO track_duration_refs
                    (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
                VALUES (?, ?, ?, ?, ?)
                """,
                (ref_id, ref_type, chosen_ms, args.source_label, now_iso),
            )
            inserted_refs += 1

            if idx % 100 == 0 or idx == len(unknown_paths):
                print(
                    f"[resolve {idx}/{len(unknown_paths)}] "
                    f"inserted_refs={inserted_refs} no_ids={skipped_no_ids} "
                    f"ambiguous_or_empty={skipped_ambiguous} guarded={skipped_guard}"
                )

        # Recompute statuses for all playlist rows present in DB.
        recomputed = 0
        for idx, path in enumerate(paths, start=1):
            row = by_path.get(path)
            if row is None:
                continue

            meta = _safe_json(row["metadata_json"])
            measured_ms = int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None
            beatport_id = _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            if not beatport_id and row["beatport_id"]:
                beatport_id = str(row["beatport_id"]).strip()

            isrc_tokens = _normalize_isrc_tokens(
                meta.get("isrc") or meta.get("ISRC") or meta.get("tsrc") or meta.get("TSRC")
            )
            if row["canonical_isrc"]:
                isrc_tokens = _normalize_isrc_tokens(isrc_tokens + [str(row["canonical_isrc"])])

            ref_ms, ref_source, ref_track_id = _lookup_ref(conn, beatport_id, isrc_tokens)
            delta_ms = measured_ms - ref_ms if (measured_ms is not None and ref_ms is not None) else None
            if delta_ms is None:
                status = "unknown"
            else:
                abs_delta = abs(delta_ms)
                if abs_delta <= ok_max_ms:
                    status = "ok"
                elif abs_delta <= warn_max_ms:
                    status = "warn"
                else:
                    status = "fail"

            conn.execute(
                """
                UPDATE files
                SET duration_ref_ms = ?,
                    duration_ref_source = ?,
                    duration_ref_track_id = ?,
                    duration_ref_updated_at = ?,
                    duration_delta_ms = ?,
                    duration_status = ?,
                    duration_check_version = ?
                WHERE path = ?
                """,
                (
                    ref_ms,
                    ref_source,
                    ref_track_id,
                    now_iso,
                    delta_ms,
                    status,
                    duration_version,
                    path,
                ),
            )
            recomputed += 1
            if idx % 200 == 0 or idx == len(paths):
                print(f"[recompute {idx}/{len(paths)}] updated={recomputed}")

        conn.commit()
        print("Done.")
        print(f"  inserted_refs:      {inserted_refs}")
        print(f"  skipped_no_ids:     {skipped_no_ids}")
        print(f"  skipped_ambiguous:  {skipped_ambiguous}")
        print(f"  skipped_guard:      {skipped_guard}")
        print(f"  recomputed_rows:    {recomputed}")
        return 0
    finally:
        beatport.close()
        spotify.close()
        tidal.close()
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
