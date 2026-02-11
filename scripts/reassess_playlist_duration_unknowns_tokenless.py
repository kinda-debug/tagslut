#!/usr/bin/env python3
"""Resolve unknown playlist duration rows without auth tokens.

Strategy:
1) Parse unknown rows for one workbook playlist (Path column).
2) Resolve references using token-free sources in strict order:
   a) Beatport track page by known beatport track id.
   b) Beatport public track search by artist/title (+ optional ISRC exact).
   c) Deezer track lookup by ISRC (guarded by title/artist similarity + delta).
3) Upsert `track_duration_refs`.
4) Recompute duration status for all playlist rows.
5) Emit a resolver CSV for auditability.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import sqlite3
import time
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Iterable
from urllib.parse import quote_plus

import httpx
from openpyxl import load_workbook

try:
    from rapidfuzz import fuzz
except Exception:  # pragma: no cover - fallback if rapidfuzz is unavailable
    from difflib import SequenceMatcher

    class _FuzzFallback:
        @staticmethod
        def ratio(a: str, b: str) -> float:
            return SequenceMatcher(None, a, b).ratio() * 100.0

        @staticmethod
        def partial_ratio(a: str, b: str) -> float:
            return SequenceMatcher(None, a, b).ratio() * 100.0

        @staticmethod
        def token_set_ratio(a: str, b: str) -> float:
            return SequenceMatcher(None, a, b).ratio() * 100.0

        @staticmethod
        def token_sort_ratio(a: str, b: str) -> float:
            return SequenceMatcher(None, a, b).ratio() * 100.0

    fuzz = _FuzzFallback()


ISRC_RE = re.compile(r"^[A-Z]{2}[A-Z0-9]{3}\d{7}$")
ISRC_SPLIT_RE = re.compile(r"[;,/\\]|\s+")
UUID_RE = re.compile(
    r"^[0-9a-f]{8}-[0-9a-f]{4}-[1-5][0-9a-f]{3}-[89ab][0-9a-f]{3}-[0-9a-f]{12}$",
    re.IGNORECASE,
)
NON_ALNUM_RE = re.compile(r"[^a-z0-9]+")
SPACE_RE = re.compile(r"\s+")
NEXT_DATA_RE = re.compile(
    r'<script id="__NEXT_DATA__"[^>]*>(.*?)</script>',
    re.IGNORECASE | re.DOTALL,
)


@dataclass
class UnknownRow:
    path: str
    measured_ms: int | None
    metadata_json: str | None
    beatport_id: str | None
    title: str
    artist: str
    isrc_tokens: list[str]
    mb_recording_ids: list[str]


@dataclass
class ResolvedRef:
    source: str
    duration_ms: int
    beatport_track_id: str | None
    isrc: str | None
    title_score: float
    artist_score: float
    abs_delta_ms: int | None
    note: str


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Resolve unknown duration statuses with token-free sources.")
    parser.add_argument("--xlsx", type=Path, required=True, help="Playlist workbook path.")
    parser.add_argument("--db", type=Path, required=True, help="SQLite DB path.")
    parser.add_argument(
        "--source-label",
        default="tokenless_authoritative_v1",
        help="Prefix for ref_source values written to track_duration_refs.",
    )
    parser.add_argument("--timeout", type=float, default=25.0, help="HTTP timeout seconds.")
    parser.add_argument("--report", type=Path, help="Resolver CSV path.")
    parser.add_argument(
        "--no-deezer",
        action="store_true",
        help="Disable Deezer ISRC fallback.",
    )
    parser.add_argument(
        "--no-musicbrainz",
        action="store_true",
        help="Disable MusicBrainz ISRC fallback.",
    )
    parser.add_argument("--min-noisrc-title-score", type=float, default=92.0)
    parser.add_argument("--min-noisrc-artist-score", type=float, default=75.0)
    parser.add_argument("--max-noisrc-delta-ms", type=int, default=4000)
    parser.add_argument("--min-isrc-title-score", type=float, default=90.0)
    parser.add_argument("--min-isrc-artist-score", type=float, default=70.0)
    parser.add_argument("--max-isrc-delta-ms", type=int, default=15000)
    return parser.parse_args()


def _safe_json(payload: str | None) -> dict:
    if not payload:
        return {}
    try:
        parsed = json.loads(payload)
    except Exception:
        return {}
    return parsed if isinstance(parsed, dict) else {}


def _normalize_text(text: str) -> str:
    lowered = (text or "").lower()
    lowered = NON_ALNUM_RE.sub(" ", lowered)
    return SPACE_RE.sub(" ", lowered).strip()


def _similarity(a: str, b: str) -> float:
    if not a or not b:
        return 0.0
    return float(
        max(
            fuzz.ratio(a, b),
            fuzz.token_set_ratio(a, b),
            fuzz.token_sort_ratio(a, b),
            fuzz.partial_ratio(a, b),
        )
    )


def _extract_tag_value(meta: dict, keys: Iterable[str]) -> str | None:
    lowered = {str(k).lower(): v for k, v in meta.items()}
    for key in keys:
        raw = lowered.get(str(key).lower())
        if raw is None:
            continue
        if isinstance(raw, list):
            if not raw:
                continue
            text = str(raw[0]).strip()
        else:
            text = str(raw).strip()
        if text:
            return text
    return None


def _normalize_isrc_tokens(value: str | None) -> list[str]:
    if not value:
        return []
    out: list[str] = []
    for token in ISRC_SPLIT_RE.split(str(value).strip().upper()):
        token = token.strip()
        if ISRC_RE.match(token):
            out.append(token)
    seen: set[str] = set()
    uniq: list[str] = []
    for token in out:
        if token in seen:
            continue
        seen.add(token)
        uniq.append(token)
    return uniq


def _normalize_uuid_tokens(value: str | None) -> list[str]:
    if not value:
        return []
    raw = str(value).strip()
    # Some tags can contain multiple IDs separated by delimiters/spaces.
    tokens = re.split(r"[;,/\s]+", raw)
    out: list[str] = []
    seen: set[str] = set()
    for token in tokens:
        token = token.strip()
        if not token:
            continue
        if not UUID_RE.match(token):
            continue
        token = token.lower()
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def _load_playlist_paths(xlsx: Path) -> list[str]:
    wb = load_workbook(xlsx, data_only=True, read_only=True)
    ws = wb["Tracks"] if "Tracks" in wb.sheetnames else wb[wb.sheetnames[0]]
    header = [c for c in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    if "Path" not in header:
        raise RuntimeError(f"'Path' column not found in {xlsx}")
    idx = header.index("Path")
    out: list[str] = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row:
            continue
        value = row[idx]
        if value:
            out.append(str(value))
    return out


def _load_scope_rows(conn: sqlite3.Connection, paths: list[str]) -> dict[str, sqlite3.Row]:
    placeholders = ",".join(["?"] * len(paths))
    rows = conn.execute(
        f"""
        SELECT path, duration_status, duration_measured_ms, metadata_json, beatport_id
        FROM files
        WHERE path IN ({placeholders})
        """,
        paths,
    ).fetchall()
    return {row["path"]: row for row in rows}


def _unknown_rows(scope_rows: dict[str, sqlite3.Row], ordered_paths: list[str]) -> list[UnknownRow]:
    out: list[UnknownRow] = []
    for path in ordered_paths:
        row = scope_rows.get(path)
        if row is None:
            continue
        if str(row["duration_status"] or "") != "unknown":
            continue
        meta = _safe_json(row["metadata_json"])
        title = _extract_tag_value(meta, ["TITLE", "title"]) or ""
        artist = _extract_tag_value(meta, ["ARTIST", "ALBUMARTIST", "artist", "albumartist"]) or ""
        beatport_id = (
            _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
            or (str(row["beatport_id"]).strip() if row["beatport_id"] else None)
        )
        isrc_tokens = _normalize_isrc_tokens(
            _extract_tag_value(meta, ["ISRC", "TSRC", "isrc", "tsrc"]) or ""
        )
        out.append(
            UnknownRow(
                path=path,
                measured_ms=int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None,
                metadata_json=row["metadata_json"],
                beatport_id=beatport_id if beatport_id else None,
                title=title,
                artist=artist,
                isrc_tokens=isrc_tokens,
                mb_recording_ids=(
                    _normalize_uuid_tokens(
                        _extract_tag_value(
                            meta,
                            [
                                "MUSICBRAINZ_TRACKID",
                                "MUSICBRAINZ_RECORDINGID",
                                "musicbrainz_trackid",
                                "musicbrainz_recordingid",
                            ],
                        )
                        or ""
                    )
                ),
            )
        )
    return out


def _lookup_ref(
    conn: sqlite3.Connection,
    beatport_id: str | None,
    isrc_tokens: list[str],
    mb_recording_ids: list[str],
) -> tuple[int | None, str | None, str | None]:
    if beatport_id:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (beatport_id,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], beatport_id
    for isrc in isrc_tokens:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (isrc,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], isrc
    for mbid in mb_recording_ids:
        row = conn.execute(
            "SELECT duration_ref_ms, ref_source FROM track_duration_refs WHERE ref_id = ?",
            (mbid,),
        ).fetchone()
        if row:
            return int(row[0]), row[1], mbid
    return None, None, None


def _upsert_ref(
    conn: sqlite3.Connection,
    ref_id: str,
    ref_type: str,
    duration_ms: int,
    source: str,
    now_iso: str,
) -> None:
    conn.execute(
        """
        INSERT OR REPLACE INTO track_duration_refs
            (ref_id, ref_type, duration_ref_ms, ref_source, ref_updated_at)
        VALUES (?, ?, ?, ?, ?)
        """,
        (ref_id, ref_type, int(duration_ms), source, now_iso),
    )


def _extract_beatport_search_tracks(payload: dict) -> list[dict]:
    queries = (
        payload.get("props", {})
        .get("pageProps", {})
        .get("dehydratedState", {})
        .get("queries", [])
    )
    if not isinstance(queries, list):
        return []
    for query in queries:
        if not isinstance(query, dict):
            continue
        state_data = query.get("state", {}).get("data")
        if isinstance(state_data, dict) and isinstance(state_data.get("data"), list):
            return state_data["data"]
    return []


def _extract_beatport_track_payload(payload: dict, track_id: str) -> dict | None:
    queries = (
        payload.get("props", {})
        .get("pageProps", {})
        .get("dehydratedState", {})
        .get("queries", [])
    )
    if not isinstance(queries, list):
        return None
    target = f"track-{track_id}"
    for query in queries:
        if not isinstance(query, dict):
            continue
        key = query.get("queryKey")
        if isinstance(key, list) and any(str(item) == target for item in key):
            data = query.get("state", {}).get("data")
            if isinstance(data, dict):
                return data
    return None


def _http_get_next_data_json(client: httpx.Client, url: str) -> dict | None:
    response = client.get(url)
    if response.status_code != 200:
        return None
    match = NEXT_DATA_RE.search(response.text)
    if not match:
        return None
    try:
        parsed = json.loads(match.group(1))
    except Exception:
        return None
    return parsed if isinstance(parsed, dict) else None


def _beatport_by_track_id(client: httpx.Client, row: UnknownRow) -> ResolvedRef | None:
    if not row.beatport_id or not row.beatport_id.isdigit():
        return None
    payload = _http_get_next_data_json(client, f"https://www.beatport.com/track/-/{row.beatport_id}")
    if not payload:
        return None
    data = _extract_beatport_track_payload(payload, row.beatport_id)
    if not data:
        return None
    length_ms = data.get("length_ms")
    if not isinstance(length_ms, int) or length_ms <= 0:
        return None
    bp_title = str(data.get("name") or data.get("track_name") or "").strip()
    artist_items = data.get("artists") if isinstance(data.get("artists"), list) else []
    bp_artist = " ".join(
        str(item.get("name") or item.get("artist_name") or "").strip()
        for item in artist_items
        if isinstance(item, dict)
    ).strip()
    bp_isrc = str(data.get("isrc") or "").strip().upper() or None
    title_score = _similarity(_normalize_text(row.title), _normalize_text(bp_title))
    artist_score = _similarity(_normalize_text(row.artist), _normalize_text(bp_artist))
    abs_delta = abs(int(row.measured_ms) - int(length_ms)) if row.measured_ms is not None else None
    return ResolvedRef(
        source="beatport_track_page",
        duration_ms=int(length_ms),
        beatport_track_id=str(row.beatport_id),
        isrc=bp_isrc if bp_isrc and ISRC_RE.match(bp_isrc) else None,
        title_score=title_score,
        artist_score=artist_score,
        abs_delta_ms=abs_delta,
        note="track_id_lookup",
    )


def _beatport_search(client: httpx.Client, row: UnknownRow, args: argparse.Namespace) -> ResolvedRef | None:
    query = " ".join(x for x in [row.artist, row.title] if x).strip() or row.title
    if not query:
        return None
    payload = _http_get_next_data_json(
        client, f"https://www.beatport.com/search/tracks?q={quote_plus(query)}"
    )
    if not payload:
        return None
    tracks = _extract_beatport_search_tracks(payload)
    if not tracks:
        return None

    local_title = _normalize_text(row.title)
    local_artist = _normalize_text(row.artist)
    local_isrcs = set(row.isrc_tokens)

    candidates: list[ResolvedRef] = []
    for track in tracks[:60]:
        if not isinstance(track, dict):
            continue
        track_id = track.get("track_id")
        length_ms = track.get("length_ms")
        if not isinstance(length_ms, int):
            length_ms = track.get("length")
        if not isinstance(track_id, int) or not isinstance(length_ms, int) or length_ms <= 0:
            continue
        title = str(track.get("track_name") or track.get("name") or "").strip()
        artists_val = track.get("artists") if isinstance(track.get("artists"), list) else []
        artist = " ".join(
            str(item.get("artist_name") or item.get("name") or "").strip()
            for item in artists_val
            if isinstance(item, dict)
        ).strip()
        isrc = str(track.get("isrc") or "").strip().upper()
        isrc = isrc if ISRC_RE.match(isrc) else None
        title_score = _similarity(local_title, _normalize_text(title))
        artist_score = _similarity(local_artist, _normalize_text(artist))
        abs_delta = abs(int(row.measured_ms) - int(length_ms)) if row.measured_ms is not None else None
        note = "search"
        candidates.append(
            ResolvedRef(
                source="beatport_search",
                duration_ms=int(length_ms),
                beatport_track_id=str(track_id),
                isrc=isrc,
                title_score=title_score,
                artist_score=artist_score,
                abs_delta_ms=abs_delta,
                note=note,
            )
        )

    if not candidates:
        return None

    # With ISRC in local tags: only accept exact ISRC matches.
    if local_isrcs:
        exact = [c for c in candidates if c.isrc and c.isrc in local_isrcs]
        if not exact:
            return None
        # Prefer closest measured duration, then text confidence.
        exact.sort(
            key=lambda c: (
                c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
                -(c.title_score + c.artist_score),
            )
        )
        best = exact[0]
        best.note = "search_isrc_exact"
        return best

    # No ISRC available: require strong title/artist + measured-duration proximity.
    text_filtered = [
        c
        for c in candidates
        if c.title_score >= args.min_noisrc_title_score and c.artist_score >= args.min_noisrc_artist_score
    ]
    if not text_filtered:
        return None
    if row.measured_ms is not None:
        text_filtered = [
            c
            for c in text_filtered
            if c.abs_delta_ms is not None and c.abs_delta_ms <= args.max_noisrc_delta_ms
        ]
        if not text_filtered:
            return None
    text_filtered.sort(
        key=lambda c: (
            c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
            -(c.title_score + c.artist_score),
        )
    )
    best = text_filtered[0]
    best.note = "search_text_duration_guarded"
    return best


def _deezer_isrc(client: httpx.Client, row: UnknownRow, args: argparse.Namespace) -> ResolvedRef | None:
    if not row.isrc_tokens:
        return None
    local_title = _normalize_text(row.title)
    local_artist = _normalize_text(row.artist)
    candidates: list[ResolvedRef] = []

    for isrc in row.isrc_tokens:
        response = client.get(f"https://api.deezer.com/track/isrc:{isrc}")
        if response.status_code != 200:
            continue
        try:
            data = response.json()
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        if "error" in data:
            continue
        resp_isrc = str(data.get("isrc") or "").strip().upper()
        if resp_isrc != isrc:
            continue
        duration_s = data.get("duration")
        if not isinstance(duration_s, int) or duration_s <= 0:
            continue
        duration_ms = int(duration_s) * 1000
        deezer_title = str(data.get("title") or "").strip()
        artist_obj = data.get("artist")
        deezer_artist = (
            str(artist_obj.get("name") or "").strip()
            if isinstance(artist_obj, dict)
            else ""
        )
        title_score = _similarity(local_title, _normalize_text(deezer_title))
        artist_score = _similarity(local_artist, _normalize_text(deezer_artist))
        abs_delta = abs(int(row.measured_ms) - duration_ms) if row.measured_ms is not None else None
        if title_score < args.min_isrc_title_score or artist_score < args.min_isrc_artist_score:
            continue
        if abs_delta is not None and abs_delta > args.max_isrc_delta_ms:
            continue
        candidates.append(
            ResolvedRef(
                source="deezer_isrc",
                duration_ms=duration_ms,
                beatport_track_id=None,
                isrc=isrc,
                title_score=title_score,
                artist_score=artist_score,
                abs_delta_ms=abs_delta,
                note="isrc_text_duration_guarded",
            )
        )

    if not candidates:
        return None
    candidates.sort(
        key=lambda c: (
            c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
            -(c.title_score + c.artist_score),
        )
    )
    return candidates[0]


def _musicbrainz_isrc(client: httpx.Client, row: UnknownRow, args: argparse.Namespace) -> ResolvedRef | None:
    if not row.isrc_tokens:
        return None
    local_title = _normalize_text(row.title)
    local_artist = _normalize_text(row.artist)
    candidates: list[ResolvedRef] = []

    for isrc in row.isrc_tokens:
        response: httpx.Response | None = None
        for attempt in range(4):
            try:
                response = client.get(
                    f"https://musicbrainz.org/ws/2/isrc/{isrc}",
                    params={"fmt": "json", "inc": "artists"},
                )
                break
            except Exception:
                response = None
                time.sleep(0.8 * (attempt + 1))
        if response is None or response.status_code != 200:
            continue
        try:
            data = response.json()
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        recordings = data.get("recordings")
        if not isinstance(recordings, list):
            continue
        for recording in recordings:
            if not isinstance(recording, dict):
                continue
            length_ms = recording.get("length")
            if not isinstance(length_ms, int) or length_ms <= 0:
                continue
            title = str(recording.get("title") or "").strip()
            artist_credit = recording.get("artist-credit")
            artists: list[str] = []
            if isinstance(artist_credit, list):
                for entry in artist_credit:
                    if not isinstance(entry, dict):
                        continue
                    if entry.get("name"):
                        artists.append(str(entry.get("name")))
                        continue
                    artist_obj = entry.get("artist")
                    if isinstance(artist_obj, dict) and artist_obj.get("name"):
                        artists.append(str(artist_obj.get("name")))
            artist = " ".join(x.strip() for x in artists if x and x.strip())
            title_score = _similarity(local_title, _normalize_text(title))
            artist_score = _similarity(local_artist, _normalize_text(artist))
            abs_delta = abs(int(row.measured_ms) - int(length_ms)) if row.measured_ms is not None else None
            if title_score < args.min_isrc_title_score or artist_score < args.min_isrc_artist_score:
                continue
            if abs_delta is not None and abs_delta > args.max_isrc_delta_ms:
                continue
            candidates.append(
                ResolvedRef(
                    source="musicbrainz_isrc",
                    duration_ms=int(length_ms),
                    beatport_track_id=None,
                    isrc=isrc,
                    title_score=title_score,
                    artist_score=artist_score,
                    abs_delta_ms=abs_delta,
                    note="isrc_text_duration_guarded",
                )
            )

    if not candidates:
        return None
    candidates.sort(
        key=lambda c: (
            c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
            -(c.title_score + c.artist_score),
        )
    )
    return candidates[0]


def _deezer_search(client: httpx.Client, row: UnknownRow, args: argparse.Namespace) -> ResolvedRef | None:
    query = " ".join(x for x in [row.artist, row.title] if x).strip() or row.title
    if not query:
        return None
    response = client.get(
        "https://api.deezer.com/search",
        params={"q": query},
    )
    if response.status_code != 200:
        return None
    try:
        data = response.json()
    except Exception:
        return None
    if not isinstance(data, dict):
        return None
    items = data.get("data")
    if not isinstance(items, list):
        return None

    local_title = _normalize_text(row.title)
    local_artist = _normalize_text(row.artist)
    local_isrcs = set(row.isrc_tokens)
    candidates: list[ResolvedRef] = []
    for item in items[:40]:
        if not isinstance(item, dict):
            continue
        duration_s = item.get("duration")
        if not isinstance(duration_s, int) or duration_s <= 0:
            continue
        duration_ms = int(duration_s) * 1000
        title = str(item.get("title") or "").strip()
        artist_obj = item.get("artist")
        artist = (
            str(artist_obj.get("name") or "").strip()
            if isinstance(artist_obj, dict)
            else ""
        )
        isrc = str(item.get("isrc") or "").strip().upper()
        isrc = isrc if ISRC_RE.match(isrc) else None
        if local_isrcs and (not isrc or isrc not in local_isrcs):
            continue
        title_score = _similarity(local_title, _normalize_text(title))
        artist_score = _similarity(local_artist, _normalize_text(artist))
        abs_delta = abs(int(row.measured_ms) - int(duration_ms)) if row.measured_ms is not None else None
        if title_score < args.min_isrc_title_score or artist_score < args.min_isrc_artist_score:
            continue
        if abs_delta is not None and abs_delta > args.max_isrc_delta_ms:
            continue
        candidates.append(
            ResolvedRef(
                source="deezer_search",
                duration_ms=duration_ms,
                beatport_track_id=None,
                isrc=isrc,
                title_score=title_score,
                artist_score=artist_score,
                abs_delta_ms=abs_delta,
                note="search_text_duration_guarded",
            )
        )
    if not candidates:
        return None
    candidates.sort(
        key=lambda c: (
            c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
            -(c.title_score + c.artist_score),
        )
    )
    return candidates[0]


def _musicbrainz_recording_id(
    client: httpx.Client, row: UnknownRow, args: argparse.Namespace
) -> ResolvedRef | None:
    if not row.mb_recording_ids:
        return None
    local_title = _normalize_text(row.title)
    local_artist = _normalize_text(row.artist)
    candidates: list[ResolvedRef] = []

    for mbid in row.mb_recording_ids:
        response: httpx.Response | None = None
        for attempt in range(4):
            try:
                response = client.get(
                    f"https://musicbrainz.org/ws/2/recording/{mbid}",
                    params={"fmt": "json", "inc": "artists"},
                )
                break
            except Exception:
                response = None
                time.sleep(0.8 * (attempt + 1))
        if response is None or response.status_code != 200:
            continue
        try:
            data = response.json()
        except Exception:
            continue
        if not isinstance(data, dict):
            continue
        length_ms = data.get("length")
        if not isinstance(length_ms, int) or length_ms <= 0:
            continue
        title = str(data.get("title") or "").strip()
        artist_credit = data.get("artist-credit")
        artists: list[str] = []
        if isinstance(artist_credit, list):
            for entry in artist_credit:
                if not isinstance(entry, dict):
                    continue
                if entry.get("name"):
                    artists.append(str(entry.get("name")))
                    continue
                artist_obj = entry.get("artist")
                if isinstance(artist_obj, dict) and artist_obj.get("name"):
                    artists.append(str(artist_obj.get("name")))
        artist = " ".join(x.strip() for x in artists if x and x.strip())
        title_score = _similarity(local_title, _normalize_text(title))
        artist_score = _similarity(local_artist, _normalize_text(artist))
        abs_delta = abs(int(row.measured_ms) - int(length_ms)) if row.measured_ms is not None else None
        # Recording MBID is a strong identity signal; keep title strict, artist softer.
        if title_score < args.min_isrc_title_score:
            continue
        candidates.append(
            ResolvedRef(
                source="musicbrainz_recording_id",
                duration_ms=int(length_ms),
                beatport_track_id=None,
                isrc=None,
                title_score=title_score,
                artist_score=artist_score,
                abs_delta_ms=abs_delta,
                note="recording_id_lookup",
            )
        )

    if not candidates:
        return None
    candidates.sort(
        key=lambda c: (
            c.abs_delta_ms if c.abs_delta_ms is not None else 10**9,
            -(c.title_score + c.artist_score),
        )
    )
    return candidates[0]


def _write_resolver_report(path: Path, rows: list[dict]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fields = [
        "path",
        "measured_ms",
        "title",
        "artist",
        "isrc_tokens",
        "beatport_id",
        "resolved",
        "resolver_source",
        "duration_ms",
        "abs_delta_ms",
        "title_score",
        "artist_score",
        "note",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def main() -> int:
    args = parse_args()
    xlsx = args.xlsx.expanduser().resolve()
    db_path = args.db.expanduser().resolve()
    if not xlsx.exists():
        raise SystemExit(f"XLSX not found: {xlsx}")
    if not db_path.exists():
        raise SystemExit(f"DB not found: {db_path}")

    report_path = args.report
    if report_path is None:
        stem = xlsx.stem.replace(" ", "_")
        report_path = xlsx.parent / f"{stem}_tokenless_unknown_resolution.csv"
    report_path = report_path.expanduser().resolve()

    paths = _load_playlist_paths(xlsx)
    if not paths:
        raise SystemExit("No paths loaded from workbook.")

    conn = sqlite3.connect(str(db_path))
    conn.row_factory = sqlite3.Row
    now_iso = datetime.now(timezone.utc).isoformat()
    ok_max_ms = 2000
    warn_max_ms = 8000
    duration_version = "duration_v1_ok2_warn8"

    resolved = 0
    unresolved = 0
    resolver_rows: list[dict] = []

    try:
        scope_rows = _load_scope_rows(conn, paths)
        unknown_rows = _unknown_rows(scope_rows, paths)
        print(f"Playlist rows: {len(paths)}")
        print(f"Unknown rows before resolver: {len(unknown_rows)}")

        client = httpx.Client(
            timeout=args.timeout,
            follow_redirects=True,
            headers={
                "User-Agent": (
                    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                    "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/132.0 Safari/537.36"
                )
            },
        )

        for idx, row in enumerate(unknown_rows, start=1):
            candidate: ResolvedRef | None = None

            # 1) Track-id direct lookup, if available.
            candidate = _beatport_by_track_id(client, row)

            # 2) Beatport text search (strictly guarded).
            if candidate is None:
                candidate = _beatport_search(client, row, args)

            # 3) Deezer ISRC fallback (strictly guarded).
            if candidate is None and not args.no_deezer:
                candidate = _deezer_isrc(client, row, args)

            # 4) MusicBrainz ISRC fallback (strictly guarded).
            if candidate is None and not args.no_musicbrainz:
                candidate = _musicbrainz_isrc(client, row, args)

            # 5) Deezer text search fallback (strictly guarded, ISRC-aware).
            if candidate is None:
                candidate = _deezer_search(client, row, args)

            # 6) MusicBrainz recording-id fallback from embedded MBID tags.
            if candidate is None and not args.no_musicbrainz:
                candidate = _musicbrainz_recording_id(client, row, args)

            if candidate is not None:
                # Write beatport track id ref when present.
                if candidate.beatport_track_id:
                    _upsert_ref(
                        conn,
                        ref_id=candidate.beatport_track_id,
                        ref_type="beatport",
                        duration_ms=candidate.duration_ms,
                        source=f"{args.source_label}:{candidate.source}",
                        now_iso=now_iso,
                    )
                    # Persist matched beatport id for rows that had no id so recompute can use it.
                    if not row.beatport_id:
                        conn.execute(
                            "UPDATE files SET beatport_id = ? WHERE path = ?",
                            (candidate.beatport_track_id, row.path),
                        )
                # Write matching local ISRC refs, or candidate ISRC if local list is empty.
                target_isrcs = [i for i in row.isrc_tokens if i and (candidate.isrc is None or i == candidate.isrc)]
                if not target_isrcs and candidate.isrc:
                    target_isrcs = [candidate.isrc]
                for isrc in target_isrcs:
                    _upsert_ref(
                        conn,
                        ref_id=isrc,
                        ref_type="isrc",
                        duration_ms=candidate.duration_ms,
                        source=f"{args.source_label}:{candidate.source}",
                        now_iso=now_iso,
                    )
                if candidate.source == "musicbrainz_recording_id":
                    for mbid in row.mb_recording_ids:
                        _upsert_ref(
                            conn,
                            ref_id=mbid,
                            ref_type="musicbrainz",
                            duration_ms=candidate.duration_ms,
                            source=f"{args.source_label}:{candidate.source}",
                            now_iso=now_iso,
                        )
                resolved += 1
                resolver_rows.append(
                    {
                        "path": row.path,
                        "measured_ms": row.measured_ms,
                        "title": row.title,
                        "artist": row.artist,
                        "isrc_tokens": ";".join(row.isrc_tokens),
                        "beatport_id": row.beatport_id or "",
                        "resolved": 1,
                        "resolver_source": candidate.source,
                        "duration_ms": candidate.duration_ms,
                        "abs_delta_ms": candidate.abs_delta_ms if candidate.abs_delta_ms is not None else "",
                        "title_score": f"{candidate.title_score:.2f}",
                        "artist_score": f"{candidate.artist_score:.2f}",
                        "note": candidate.note,
                    }
                )
            else:
                unresolved += 1
                resolver_rows.append(
                    {
                        "path": row.path,
                        "measured_ms": row.measured_ms,
                        "title": row.title,
                        "artist": row.artist,
                        "isrc_tokens": ";".join(row.isrc_tokens),
                        "beatport_id": row.beatport_id or "",
                        "resolved": 0,
                        "resolver_source": "",
                        "duration_ms": "",
                        "abs_delta_ms": "",
                        "title_score": "",
                        "artist_score": "",
                        "note": "no_strict_match",
                    }
                )

            if idx % 50 == 0 or idx == len(unknown_rows):
                print(f"[resolve {idx}/{len(unknown_rows)}] resolved={resolved} unresolved={unresolved}")

        client.close()

        # Recompute status for all playlist rows.
        scope_rows = _load_scope_rows(conn, paths)
        for idx, path in enumerate(paths, start=1):
            row = scope_rows.get(path)
            if row is None:
                continue
            meta = _safe_json(row["metadata_json"])
            beatport_id = (
                _extract_tag_value(meta, ["BEATPORT_TRACK_ID", "BP_TRACK_ID", "beatport_track_id"])
                or (str(row["beatport_id"]).strip() if row["beatport_id"] else None)
            )
            isrc_tokens = _normalize_isrc_tokens(
                _extract_tag_value(meta, ["ISRC", "TSRC", "isrc", "tsrc"]) or ""
            )
            mb_recording_ids = _normalize_uuid_tokens(
                _extract_tag_value(
                    meta,
                    [
                        "MUSICBRAINZ_TRACKID",
                        "MUSICBRAINZ_RECORDINGID",
                        "musicbrainz_trackid",
                        "musicbrainz_recordingid",
                    ],
                )
                or ""
            )
            measured = int(row["duration_measured_ms"]) if row["duration_measured_ms"] is not None else None
            ref_ms, ref_source, ref_track_id = _lookup_ref(conn, beatport_id, isrc_tokens, mb_recording_ids)
            delta = measured - ref_ms if (measured is not None and ref_ms is not None) else None
            if delta is None:
                status = "unknown"
            else:
                ad = abs(delta)
                if ad <= ok_max_ms:
                    status = "ok"
                elif ad <= warn_max_ms:
                    status = "warn"
                else:
                    status = "fail"
            conn.execute(
                """
                UPDATE files
                SET duration_ref_ms = ?,
                    duration_ref_source = ?,
                    duration_ref_track_id = ?,
                    duration_ref_updated_at = ?,
                    duration_delta_ms = ?,
                    duration_status = ?,
                    duration_check_version = ?
                WHERE path = ?
                """,
                (ref_ms, ref_source, ref_track_id, now_iso, delta, status, duration_version, path),
            )
            if idx % 200 == 0 or idx == len(paths):
                print(f"[recompute {idx}/{len(paths)}]")

        conn.commit()

        # Unknown-after count in scope.
        scope_rows = _load_scope_rows(conn, paths)
        unknown_after = sum(1 for row in scope_rows.values() if str(row["duration_status"] or "") == "unknown")
        print("Done.")
        print(f"  resolved_unknown_rows:   {resolved}")
        print(f"  unresolved_unknown_rows: {unresolved}")
        print(f"  unknown_after_recompute: {unknown_after}")

        _write_resolver_report(report_path, resolver_rows)
        print(f"  resolver_report:         {report_path}")
        return 0
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
