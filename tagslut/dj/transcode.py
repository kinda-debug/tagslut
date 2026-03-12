"""Transcoding helpers for DJ exports."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
import re
import subprocess
import os
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook

REQUIRED_COLUMNS = {
    "Album Artist",
    "Album",
    "Track#",
    "Title",
    "Track Artist(s)",
    "External Id",
    "Source",
    "Path",
}


def run_checked(command: Sequence[str], timeout: Optional[int] = None) -> subprocess.CompletedProcess[str]:
    """Run a subprocess command and capture output without raising on failure."""
    return subprocess.run(command, capture_output=True, text=True, timeout=timeout, check=False)


def normalize_text(value: Optional[object]) -> str:
    """Normalize text for comparison and dedupe keys."""
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def sanitize_component(value: Optional[object], fallback: str) -> str:
    """Sanitize a path component for filesystem output."""
    text = str(value).strip() if value is not None else ""
    if not text:
        text = fallback
    text = re.sub(r"[\\/:*?\"<>|]", "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.rstrip(". ")
    return text or fallback


MAX_REL_PATH_LEN = 240


def _truncate(text: str, max_len: int) -> str:
    if len(text) <= max_len:
        return text
    return text[:max_len].rstrip()


def _truncate_filename(name: str, max_len: int) -> str:
    if len(name) <= max_len:
        return name
    if "." in name:
        stem, ext = name.rsplit(".", 1)
        ext = f".{ext}"
        avail = max_len - len(ext)
        if avail <= 1:
            return name[:max_len]
        return _truncate(stem, avail) + ext
    return _truncate(name, max_len)


def parse_track_number(value: Optional[object]) -> Optional[int]:
    """Parse an optional track number field into an integer."""
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "/" in text:
        text = text.split("/", 1)[0]
    try:
        return int(float(text))
    except ValueError:
        return None


@dataclass
class TrackRow:
    """Track row used for DJ transcode manifests."""

    row_num: int
    album_artist: str
    album: str
    track_number: Optional[int]
    title: str
    track_artist: str
    external_id: str
    source: str
    source_path: Path
    dedupe_key: Tuple[str, ...]
    output_path: Optional[Path] = None
    canonical_key: Optional[str] = None
    dj_set_role: Optional[str] = None


def make_dedupe_key(track: TrackRow) -> Tuple[str, ...]:
    """Build a deterministic dedupe key for a track row."""
    ext = normalize_text(track.external_id)
    if ext:
        return ("id", ext)
    artist = normalize_text(track.track_artist) or normalize_text(track.album_artist)
    title = normalize_text(track.title)
    return ("meta", artist, title)


def load_tracks(
    xlsx_path: Path,
    sheet_name: Optional[str],
    *,
    check_paths: bool = True,
) -> Tuple[List[TrackRow], List[Dict[str, object]], List[str]]:
    """Load tracks from XLSX for DJ export."""
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): i for i, h in enumerate(headers) if h is not None}

    missing_cols = sorted(REQUIRED_COLUMNS - set(header_to_idx.keys()))
    if missing_cols:
        raise RuntimeError(f"Missing required columns in worksheet '{ws.title}': {missing_cols}")

    tracks: List[TrackRow] = []
    dropped_missing_path: List[Dict[str, object]] = []

    for row_num in range(2, ws.max_row + 1):
        def get(col: str) -> Optional[object]:
            return ws.cell(row_num, header_to_idx[col] + 1).value  # type: ignore  # TODO: mypy-strict

        raw_path = get("Path")
        if raw_path is None or str(raw_path).strip() == "":
            dropped_missing_path.append({"row_num": row_num, "reason": "empty_path", "path": ""})
            continue

        source_path = Path(str(raw_path))
        if check_paths and not source_path.exists():
            dropped_missing_path.append(
                {"row_num": row_num, "reason": "missing_on_disk", "path": str(source_path)}
            )
            continue

        track = TrackRow(
            row_num=row_num,
            album_artist=str(get("Album Artist") or ""),
            album=str(get("Album") or ""),
            track_number=parse_track_number(get("Track#")),
            title=str(get("Title") or ""),
            track_artist=str(get("Track Artist(s)") or ""),
            external_id=str(get("External Id") or ""),
            source=str(get("Source") or ""),
            source_path=source_path,
            dedupe_key=("",),
        )
        track.dedupe_key = make_dedupe_key(track)
        tracks.append(track)

    return tracks, dropped_missing_path, headers


def build_output_path(output_root: Path, track: TrackRow) -> Path:
    """Construct the output path for a track."""
    artist_dir = sanitize_component(track.album_artist or track.track_artist, "Unknown Artist")
    if track.album and track.album.strip():
        album_dir = sanitize_component(track.album, "Unknown Album")
    else:
        album_dir = sanitize_component(track.source_path.parent.name, "Unknown Album")
    title = sanitize_component(track.title, track.source_path.stem)

    if track.track_number is not None:
        file_name = f"{track.track_number:02d} {title}.mp3"
    else:
        file_name = f"{title}.mp3"

    output = output_root / artist_dir / album_dir / file_name
    rel = str(output.relative_to(output_root))
    if len(rel) <= MAX_REL_PATH_LEN:
        return output

    artist_dir = _truncate(artist_dir, 60)
    album_dir = _truncate(album_dir, 60)
    file_name = _truncate_filename(file_name, 120)
    output = output_root / artist_dir / album_dir / file_name
    rel = str(output.relative_to(output_root))
    if len(rel) <= MAX_REL_PATH_LEN:
        return output

    album_dir = "Misc"
    avail = max(32, MAX_REL_PATH_LEN - len(artist_dir) - len(album_dir) - 2)
    file_name = _truncate_filename(file_name, avail)
    output = output_root / artist_dir / album_dir / file_name
    rel = str(output.relative_to(output_root))
    if len(rel) <= MAX_REL_PATH_LEN:
        return output

    file_name = _truncate_filename(file_name, MAX_REL_PATH_LEN)
    return output_root / file_name


def dedupe_tracks(tracks: Iterable[TrackRow]) -> Tuple[List[TrackRow], List[Dict[str, object]]]:
    """Deduplicate tracks by external id or artist/title identity."""
    kept: Dict[Tuple[str, ...], TrackRow] = {}
    duplicates: List[Dict[str, object]] = []

    def source_priority(value: str) -> int:
        v = normalize_text(value)
        if v == "local":
            return 0
        if v == "tidal":
            return 1
        if v == "streaming":
            return 2
        return 3

    for track in tracks:
        existing = kept.get(track.dedupe_key)
        if existing is None:
            kept[track.dedupe_key] = track
            continue

        current_rank = (source_priority(track.source), len(track.source_path.name), track.row_num)
        existing_rank = (source_priority(existing.source), len(
            existing.source_path.name), existing.row_num)

        if current_rank < existing_rank:
            kept[track.dedupe_key] = track
            duplicates.append(
                {
                    "dedupe_key": " | ".join(track.dedupe_key),
                    "kept_row": track.row_num,
                    "kept_path": str(track.source_path),
                    "dropped_row": existing.row_num,
                    "dropped_path": str(existing.source_path),
                    "reason": "better_source_or_filename",
                }
            )
        else:
            duplicates.append(
                {
                    "dedupe_key": " | ".join(track.dedupe_key),
                    "kept_row": existing.row_num,
                    "kept_path": str(existing.source_path),
                    "dropped_row": track.row_num,
                    "dropped_path": str(track.source_path),
                    "reason": "duplicate_identity",
                }
            )

    deduped = list(sorted(kept.values(), key=lambda t: t.row_num))
    return deduped, duplicates


def assign_output_paths(tracks: List[TrackRow], output_root: Path) -> None:
    """Assign output paths in-place, resolving naming collisions."""
    used_paths: Dict[Path, int] = {}
    for track in tracks:
        proposed = build_output_path(output_root, track)
        count = used_paths.get(proposed, 0)
        if count > 0:
            stem = proposed.stem
            proposed = proposed.with_name(f"{stem}__{count + 1}{proposed.suffix}")
        used_paths[proposed] = count + 1
        track.output_path = proposed


def transcode_one(
    track: TrackRow,
    overwrite: bool,
    timeout_s: int | None = None,
) -> Tuple[str, TrackRow, str, str]:
    """Transcode a single track to MP3 320 CBR.

    Returns a tuple of (status, track, error_message).
    """
    if track.output_path is None:
        return ("error", track, "missing_output_path", "")

    out = track.output_path
    out.parent.mkdir(parents=True, exist_ok=True)

    if out.exists() and not overwrite:
        return ("skipped_existing", track, "output_exists", "")

    cmd = [
        "ffmpeg",
        "-nostdin",
        "-v",
        "error",
        "-y",
        "-i",
        str(track.source_path),
        "-map_metadata",
        "0",
        "-map",
        "a:0",
        "-id3v2_version",
        "3",
        "-codec:a",
        "libmp3lame",
        "-b:a",
        "320k",
        "-minrate",
        "320k",
        "-maxrate",
        "320k",
        "-bufsize",
        "640k",
        "-write_xing",
        "0",
        str(out),
    ]
    timeout = timeout_s
    if timeout is None:
        timeout_env = os.environ.get("DJ_TRANSCODE_TIMEOUT_S")
        if timeout_env:
            try:
                timeout_val = int(timeout_env)
                if timeout_val > 0:
                    timeout = timeout_val
            except ValueError:
                timeout = None
    try:
        cp = run_checked(cmd, timeout=timeout)
    except subprocess.TimeoutExpired:
        return ("error", track, f"ffmpeg_timeout_{timeout}s" if timeout else "ffmpeg_timeout", "timeout")
    if cp.returncode == 0 and out.exists() and out.stat().st_size > 0:
        return ("ok", track, "", "")

    stderr = (cp.stderr or "").strip()
    err = stderr or "ffmpeg_failed"
    return ("error", track, err, stderr)
