#!/usr/bin/env python3
"""Transcode Excel backlog FLACs into MP3_LIBRARY + DJ_LIBRARY copies."""
from __future__ import annotations

import argparse
import logging
import os
import sqlite3
import sys
import time
from pathlib import Path
from typing import Iterable

try:
    from openpyxl import load_workbook
except ImportError as exc:
    raise SystemExit(
        "openpyxl is required to read Excel backlogs. Install it with `pip install openpyxl`.\n"
        + str(exc)
    )

from tagslut.exec.mp3_build import (
    build_full_tag_mp3_assets_from_flac_paths,
    build_dj_copies_from_full_tag_mp3_assets,
)
from tagslut.utils.db import resolve_cli_env_db_path

BACKLOG_DEFAULT = "/Users/georgeskhawam/Music/20260312_27.xlsx"
DJ_PROFILE = "dj_copy_320_cbr"
FULL_PROFILE = "mp3_asset_320_cbr_full"


def _unique_preserve_order(values: Iterable[Path]) -> list[Path]:
    seen: set[Path] = set()
    result: list[Path] = []
    for value in values:
        if value not in seen:
            seen.add(value)
            result.append(value)
    return result


def _best_asset_info(conn: sqlite3.Connection, flac_path: Path) -> tuple[int | None, int | None]:
    row = conn.execute(
        """
        SELECT af.id, al.identity_id, al.asset_id
        FROM asset_file af
        LEFT JOIN asset_link al
          ON al.asset_id = af.id
         AND (al.active IS NULL OR al.active = 1)
        WHERE af.path = ?
        ORDER BY al.confidence DESC, al.id ASC
        LIMIT 1
        """,
        (str(flac_path),),
    ).fetchone()
    if not row:
        return None, None
    _, identity_id, asset_id = row
    return identity_id, asset_id


def _has_asset(conn: sqlite3.Connection, identity_id: int, asset_id: int, profile: str) -> bool:
    row = conn.execute(
        """
        SELECT 1 FROM mp3_asset
        WHERE identity_id = ?
          AND asset_id = ?
          AND profile = ?
          AND status = 'verified'
        LIMIT 1
        """,
        (identity_id, asset_id, profile),
    ).fetchone()
    return bool(row)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument(
        "--workbook",
        default=os.environ.get("BACKLOG_WORKBOOK", BACKLOG_DEFAULT),
        help=f"Path to backlog Excel workbook (default: {BACKLOG_DEFAULT})",
    )
    parser.add_argument(
        "--column",
        default="Path",
        help="Column header containing the FLAC path",
    )
    parser.add_argument(
        "--dry-run",
        action="store_true",
        help="Only log what would be transcoded/copied, do not write files.",
    )
    parser.add_argument(
        "--verbose",
        action="store_true",
        help="Log each FLAC as it is processed.",
    )
    parser.add_argument(
        "--log",
        default=None,
        help="Path to optional logfile (appends).",
    )
    parser.add_argument(
        "--limit",
        type=int,
        default=None,
        help="Process at most N rows from the backlog (useful for testing).",
    )
    args = parser.parse_args()

    start_ts = time.time()
    handlers = [logging.StreamHandler(sys.stdout)]
    if args.log:
        handlers.append(logging.FileHandler(args.log))
    logging.basicConfig(
        level=logging.DEBUG if args.verbose else logging.INFO,
        format="%(asctime)s %(levelname)s %(message)s",
        handlers=handlers,
    )
    logger = logging.getLogger("backlog")
    logger.info("Starting backlog run dry-run=%s limit=%s", args.dry_run, args.limit)

    workbook_path = Path(args.workbook).expanduser()
    if not workbook_path.exists():
        raise SystemExit(f"Backlog workbook not found: {workbook_path}")

    mp3_root = Path(os.environ.get("MP3_LIBRARY", "")).expanduser()
    dj_root = Path(os.environ.get("DJ_LIBRARY", "")).expanduser()
    if not mp3_root or not dj_root:
        raise SystemExit("MP3_LIBRARY and DJ_LIBRARY must be set in the environment")

    db_path = resolve_cli_env_db_path(None, purpose="write", source_label="--db").path
    conn = sqlite3.connect(str(db_path))

    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = wb.active
    header = [
        str(cell).strip() if cell is not None else ""
        for cell in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))
    ]
    try:
        path_idx = header.index(args.column)
    except ValueError:
        raise SystemExit(f"Column '{args.column}' not found in {workbook_path}")

    backlog: list[Path] = []
    missing_files: list[Path] = []
    missing_db: list[Path] = []
    rows_seen = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if args.limit is not None and rows_seen >= args.limit:
            break
        raw = row[path_idx] if len(row) > path_idx else None
        if not raw:
            continue
        flac_path = Path(str(raw).strip()).expanduser()
        rows_seen += 1
        if not flac_path.exists():
            missing_files.append(flac_path)
            continue
        backlog.append(flac_path)
        if args.verbose:
            logger.debug("Found backlog flac: %s", flac_path)

    if not backlog:
        print("No FLAC paths found in backlog workbook.")
        return

    candidates: list[Path] = []
    full_tag_missing: list[Path] = []
    for flac in backlog:
        identity_id, asset_id = _best_asset_info(conn, flac)
        if identity_id is None or asset_id is None:
            missing_db.append(flac)
            continue
        if _has_asset(conn, identity_id, asset_id, DJ_PROFILE):
            continue
        if args.verbose:
            logger.debug("Candidate %s identity=%s", flac, identity_id)
        candidates.append(flac)
        if not _has_asset(conn, identity_id, asset_id, FULL_PROFILE):
            full_tag_missing.append(flac)

    if missing_files:
        print(f"{len(missing_files)} workbook rows pointed to missing FLAC files (see below):")
        for path in missing_files[:20]:
            print("  MISSING", path)
        if len(missing_files) > 20:
            print(f"  ... +{len(missing_files) - 20} more")

    if missing_db:
        print(f"{len(missing_db)} workbook rows lack catalog entries in the DB (see below):")
        for path in missing_db[:20]:
            print("  NO_DB", path)
        if len(missing_db) > 20:
            print(f"  ... +{len(missing_db) - 20} more")

    candidates = _unique_preserve_order(candidates)
    full_tag_missing = _unique_preserve_order(full_tag_missing)

    if not candidates:
        logger.info("All backlog FLACs already have DJ copies. Nothing to do.")
        print("All backlog FLACs already have DJ copies. Nothing to do.")
        return

    result_full = None
    if full_tag_missing:
        logger.info("Scheduling %d new full-tag MP3 builds", len(full_tag_missing))
        print(f"{len(full_tag_missing)} FLACs are missing MP3_LIBRARY (full-tag) assets.")
        result_full = build_full_tag_mp3_assets_from_flac_paths(
            conn,
            flac_paths=full_tag_missing,
            mp3_root=mp3_root,
            dry_run=args.dry_run,
        )
        print(result_full.summary())
        if result_full.errors:
            print("Full-tag errors:")
            for err in result_full.errors:
                print(" ", err)
                logger.error("Full-tag error: %s", err)

    logger.info("Building DJ copies for %d backlog FLACs (dry-run=%s).", len(candidates), args.dry_run)
    print(f"Building DJ copies for {len(candidates)} backlog FLACs (dry-run={args.dry_run}).")
    result_dj = build_dj_copies_from_full_tag_mp3_assets(
        conn,
        flac_paths=candidates,
        dj_root=dj_root,
        dry_run=args.dry_run,
    )
    print(result_dj.summary())
    if result_dj.errors:
        print("DJ copy errors:")
        for err in result_dj.errors:
            print(" ", err)
            logger.error("DJ copy error: %s", err)

    elapsed = time.time() - start_ts
    logger.info("Backlog run complete (%0.1f s)", elapsed)
    print("\nBacklog summary:")
    print(f"  workbook rows scanned: {rows_seen}")
    print(f"  missing files: {len(missing_files)}")
    print(f"  missing DB entries: {len(missing_db)}")
    print(f"  DJ copies produced (candidates): {len(candidates)}")
    print(f"  full-tag builds scheduled: {len(full_tag_missing)}")
    print(f"  dry-run mode: {args.dry_run}")
    if result_full:
        print(f"  full-tag summary: {result_full.summary()}")
        logger.info("Full-tag summary: %s", result_full.summary())
    print(f"  DJ copy summary: {result_dj.summary()}")
    logger.info("DJ copy summary: %s", result_dj.summary())


if __name__ == "__main__":
    main()
