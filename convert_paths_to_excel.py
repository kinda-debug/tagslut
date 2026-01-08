#!/usr/bin/env python3
"""Convert file paths to Excel with each path component in a separate column."""

import openpyxl
from pathlib import Path


def convert_paths_to_excel(input_file, output_file):
    """
    Read paths from input file and write to Excel with each path component in a separate column.
    
    Args:
        input_file: Path to the text file containing file paths (one per line)
        output_file: Path to the output Excel file
    """
    # Create a new workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "File Paths"
    
    # Read the input file
    with open(input_file, 'r', encoding='utf-8') as f:
        paths = f.readlines()
    
    # Process each path
    for row_idx, path in enumerate(paths, start=1):
        path = path.strip()
        if not path:
            continue
            
        # Split the path by '/'
        # Remove empty strings from leading '/'
        parts = [p for p in path.split('/') if p]
        
        # Write each part to a separate column
        for col_idx, part in enumerate(parts, start=1):
            ws.cell(row=row_idx, column=col_idx, value=part)
    
    # Save the workbook
    wb.save(output_file)
    print(f"✓ Converted {len(paths)} paths to {output_file}")
    print(f"✓ Maximum path depth: {ws.max_column} levels")


if __name__ == "__main__":
    input_file = "uh.txt"
    output_file = "uh_paths.xlsx"
    
    convert_paths_to_excel(input_file, output_file)
