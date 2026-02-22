#!/usr/bin/env python3
"""
    Build a DJ_YES manifest from XLSX, removing duplicates, and transcode tracks to MP3 LAME 320 CBR.

Default workflow:
1) Read `/Users/georgeskhawam/Desktop/DJ_YES.xlsx`
2) Keep rows with a valid on-disk `Path`
3) Deduplicate by:
   - normalized `External Id` when present
   - otherwise normalized (`Track Artist(s)` or `Album Artist`) + `Title`
4) Estimate runtime and output volume
5) Optionally transcode into `/Volumes/MUSIC/DJ_YES`

Usage examples:
  python scripts/dj_yes_transcode.py --estimate-only
  python scripts/dj_yes_transcode.py --dry-run
  python scripts/dj_yes_transcode.py --jobs 6
"""

from __future__ import annotations

import argparse
import csv
import json
import os
import random
import re
import shutil
import subprocess
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple

from openpyxl import load_workbook
from mutagen import File as MutagenFile

REQUIRED_COLUMNS = {
    "Album Artist",
    "Album",
    "Track#",
    "Title",
    "Track Artist(s)",
    "External Id",
    "Source",
    "Path",
}


def normalize_text(value: Optional[object]) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    text = re.sub(r"\s+", " ", text)
    return text


def sanitize_component(value: Optional[object], fallback: str) -> str:
    text = str(value).strip() if value is not None else ""
    if not text:
        text = fallback
    text = re.sub(r"[\\/:*?\"<>|]", "_", text)
    text = re.sub(r"\s+", " ", text).strip()
    text = text.rstrip(". ")
    return text or fallback


def parse_track_number(value: Optional[object]) -> Optional[int]:
    if value is None:
        return None
    text = str(value).strip()
    if not text:
        return None
    if "/" in text:
        text = text.split("/", 1)[0]
    try:
        return int(float(text))
    except ValueError:
        return None


def run_checked(command: Sequence[str], timeout: Optional[int] = None) -> subprocess.CompletedProcess:
    return subprocess.run(command, capture_output=True, text=True, timeout=timeout, check=False)


@dataclass
class TrackRow:
    row_num: int
    album_artist: str
    album: str
    track_number: Optional[int]
    title: str
    track_artist: str
    external_id: str
    source: str
    source_path: Path
    dedupe_key: Tuple[str, ...]
    output_path: Optional[Path] = None


def make_dedupe_key(track: TrackRow) -> Tuple[str, ...]:
    ext = normalize_text(track.external_id)
    if ext:
        return ("id", ext)
    artist = normalize_text(track.track_artist) or normalize_text(track.album_artist)
    title = normalize_text(track.title)
    return ("meta", artist, title)


def build_output_path(output_root: Path, track: TrackRow) -> Path:
    artist_dir = sanitize_component(track.album_artist or track.track_artist, "Unknown Artist")
    album_dir = sanitize_component(track.album, "Unknown Album")
    title = sanitize_component(track.title, track.source_path.stem)

    if track.track_number is not None:
        file_name = f"{track.track_number:02d} {title}.mp3"
    else:
        file_name = f"{title}.mp3"

    return output_root / artist_dir / album_dir / file_name


def check_ffmpeg() -> None:
    ffmpeg_path = shutil.which("ffmpeg")
    if not ffmpeg_path:
        raise RuntimeError("ffmpeg is required but not found on PATH.")

    probe = run_checked(["ffmpeg", "-version"])
    if probe.returncode != 0:
        raise RuntimeError(f"ffmpeg exists but is not runnable: {probe.stderr.strip()}")


def load_tracks(xlsx_path: Path, sheet_name: Optional[str]) -> Tuple[List[TrackRow], List[Dict[str, object]], List[str]]:
    wb = load_workbook(xlsx_path, data_only=True, read_only=True)
    ws = wb[sheet_name] if sheet_name else wb[wb.sheetnames[0]]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    header_to_idx = {str(h): i for i, h in enumerate(headers) if h is not None}

    missing_cols = sorted(REQUIRED_COLUMNS - set(header_to_idx.keys()))
    if missing_cols:
        raise RuntimeError(f"Missing required columns in worksheet '{ws.title}': {missing_cols}")

    tracks: List[TrackRow] = []
    dropped_missing_path: List[Dict[str, object]] = []

    for row_num in range(2, ws.max_row + 1):
        def get(col: str) -> Optional[object]:
            return ws.cell(row_num, header_to_idx[col] + 1).value

        raw_path = get("Path")
        if raw_path is None or str(raw_path).strip() == "":
            dropped_missing_path.append({"row_num": row_num, "reason": "empty_path", "path": ""})
            continue

        source_path = Path(str(raw_path))
        if not source_path.exists():
            dropped_missing_path.append({"row_num": row_num, "reason": "missing_on_disk", "path": str(source_path)})
            continue

        track = TrackRow(
            row_num=row_num,
            album_artist=str(get("Album Artist") or ""),
            album=str(get("Album") or ""),
            track_number=parse_track_number(get("Track#")),
            title=str(get("Title") or ""),
            track_artist=str(get("Track Artist(s)") or ""),
            external_id=str(get("External Id") or ""),
            source=str(get("Source") or ""),
            source_path=source_path,
            dedupe_key=("",),
        )
        track.dedupe_key = make_dedupe_key(track)
        tracks.append(track)

    return tracks, dropped_missing_path, headers


def dedupe_tracks(tracks: Iterable[TrackRow]) -> Tuple[List[TrackRow], List[Dict[str, object]]]:
    kept: Dict[Tuple[str, ...], TrackRow] = {}
    duplicates: List[Dict[str, object]] = []

    def source_priority(value: str) -> int:
        v = normalize_text(value)
        if v == "local":
            return 0
        if v == "tidal":
            return 1
        if v == "streaming":
            return 2
        return 3

    for track in tracks:
        existing = kept.get(track.dedupe_key)
        if existing is None:
            kept[track.dedupe_key] = track
            continue

        current_rank = (source_priority(track.source), len(track.source_path.name), track.row_num)
        existing_rank = (source_priority(existing.source), len(existing.source_path.name), existing.row_num)

        if current_rank < existing_rank:
            kept[track.dedupe_key] = track
            duplicates.append(
                {
                    "dedupe_key": " | ".join(track.dedupe_key),
                    "kept_row": track.row_num,
                    "kept_path": str(track.source_path),
                    "dropped_row": existing.row_num,
                    "dropped_path": str(existing.source_path),
                    "reason": "better_source_or_filename",
                }
            )
        else:
            duplicates.append(
                {
                    "dedupe_key": " | ".join(track.dedupe_key),
                    "kept_row": existing.row_num,
                    "kept_path": str(existing.source_path),
                    "dropped_row": track.row_num,
                    "dropped_path": str(track.source_path),
                    "reason": "duplicate_identity",
                }
            )

    deduped = list(sorted(kept.values(), key=lambda t: t.row_num))
    return deduped, duplicates


def assign_output_paths(tracks: List[TrackRow], output_root: Path) -> None:
    used_paths: Dict[Path, int] = {}
    for track in tracks:
        proposed = build_output_path(output_root, track)
        count = used_paths.get(proposed, 0)
        if count > 0:
            stem = proposed.stem
            proposed = proposed.with_name(f"{stem}__{count + 1}{proposed.suffix}")
        used_paths[proposed] = count + 1
        track.output_path = proposed


def ffprobe_duration_seconds(path: Path, timeout_sec: int = 8) -> Optional[float]:
    cmd = [
        "ffprobe",
        "-v",
        "error",
        "-show_entries",
        "format=duration",
        "-of",
        "default=noprint_wrappers=1:nokey=1",
        str(path),
    ]
    try:
        cp = run_checked(cmd, timeout=timeout_sec)
    except subprocess.TimeoutExpired:
        return None
    if cp.returncode != 0:
        return None
    text = cp.stdout.strip()
    try:
        value = float(text)
    except ValueError:
        return None
    if value <= 0:
        return None
    return value


def fast_duration_seconds(path: Path) -> Optional[float]:
    try:
        audio = MutagenFile(path)
    except Exception:
        audio = None
    if audio is not None and hasattr(audio, "info") and hasattr(audio.info, "length"):
        try:
            d = float(audio.info.length)
            if d > 0:
                return d
        except Exception:
            pass
    # Fallback for problematic files.
    return ffprobe_duration_seconds(path, timeout_sec=8)


def estimate_duration_and_size(tracks: List[TrackRow], sample_size: int, sample_seed: int) -> Dict[str, object]:
    n_tracks = len(tracks)
    if n_tracks == 0:
        return {
            "sample_size": 0,
            "sample_success": 0,
            "avg_duration_sec": 0.0,
            "total_duration_sec_est": 0.0,
            "total_duration_h_est": 0.0,
            "output_bytes_est": 0,
            "output_gb_est": 0.0,
        }

    rng = random.Random(sample_seed)
    sample_n = min(sample_size, n_tracks)
    sample_tracks = rng.sample(tracks, sample_n)

    durations: List[float] = []
    for track in sample_tracks:
        d = fast_duration_seconds(track.source_path)
        if d is not None:
            durations.append(d)

    known_count = len(durations)
    if known_count == 0:
        avg_duration = 240.0
        total_duration = avg_duration * n_tracks
    else:
        avg_duration = sum(durations) / known_count
        total_duration = avg_duration * n_tracks
    # 320 kbps CBR -> 40,000 bytes/s (+ small metadata overhead per file).
    output_bytes = int(total_duration * 40000 + n_tracks * 4096)

    return {
        "sample_size": sample_n,
        "sample_success": known_count,
        "avg_duration_sec": avg_duration,
        "total_duration_sec_est": total_duration,
        "total_duration_h_est": total_duration / 3600,
        "output_bytes_est": output_bytes,
        "output_gb_est": output_bytes / (1024**3),
    }


def benchmark_realtime_factor(
    tracks: List[TrackRow],
    sample_size: int,
    sample_seed: int,
    timeout_sec: int,
) -> Dict[str, object]:
    if not tracks or sample_size <= 0:
        return {
            "sample_size": 0,
            "sample_success": 0,
            "avg_realtime_x": 0.0,
            "min_realtime_x": 0.0,
            "max_realtime_x": 0.0,
        }

    rng = random.Random(sample_seed + 17)
    sample_n = min(sample_size, len(tracks))
    sample_tracks = rng.sample(tracks, sample_n)

    speeds: List[float] = []
    for track in sample_tracks:
        duration = ffprobe_duration_seconds(track.source_path)
        if duration is None or duration <= 0:
            continue

        cmd = [
            "ffmpeg",
            "-nostdin",
            "-v",
            "error",
            "-y",
            "-i",
            str(track.source_path),
            "-map_metadata",
            "0",
            "-map",
            "a:0",
            "-codec:a",
            "libmp3lame",
            "-b:a",
            "320k",
            "-minrate",
            "320k",
            "-maxrate",
            "320k",
            "-bufsize",
            "640k",
            "-write_xing",
            "0",
            "-f",
            "mp3",
            "/dev/null",
        ]
        start = time.perf_counter()
        try:
            cp = run_checked(cmd, timeout=timeout_sec)
        except subprocess.TimeoutExpired:
            continue
        elapsed = time.perf_counter() - start
        if cp.returncode != 0 or elapsed <= 0:
            continue
        speeds.append(duration / elapsed)

    if not speeds:
        return {
            "sample_size": sample_n,
            "sample_success": 0,
            "avg_realtime_x": 0.0,
            "min_realtime_x": 0.0,
            "max_realtime_x": 0.0,
        }

    return {
        "sample_size": sample_n,
        "sample_success": len(speeds),
        "avg_realtime_x": sum(speeds) / len(speeds),
        "min_realtime_x": min(speeds),
        "max_realtime_x": max(speeds),
    }


def estimate_runtime_hours(total_duration_sec_est: float, avg_realtime_x: float, jobs: int) -> Dict[str, float]:
    if total_duration_sec_est <= 0:
        return {"best": 0.0, "likely": 0.0, "worst": 0.0}

    if avg_realtime_x <= 0:
        # Safe fallback if benchmark failed.
        effective_rt = max(1.5, jobs * 1.0)
    else:
        # Parallel efficiency rarely scales linearly for file IO + codec workloads.
        effective_rt = avg_realtime_x * max(1, jobs) * 0.78

    likely_sec = total_duration_sec_est / effective_rt
    best_sec = likely_sec * 0.8
    worst_sec = likely_sec * 1.35
    return {"best": best_sec / 3600, "likely": likely_sec / 3600, "worst": worst_sec / 3600}


def write_csv(path: Path, rows: List[Dict[str, object]], fieldnames: Sequence[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def transcode_one(track: TrackRow, overwrite: bool) -> Tuple[str, TrackRow, str]:
    assert track.output_path is not None
    out = track.output_path
    out.parent.mkdir(parents=True, exist_ok=True)

    if out.exists() and not overwrite:
        return ("skipped_existing", track, "output_exists")

    cmd = [
        "ffmpeg",
        "-nostdin",
        "-v",
        "error",
        "-y",
        "-i",
        str(track.source_path),
        "-map_metadata",
        "0",
        "-map",
        "a:0",
        "-id3v2_version",
        "3",
        "-codec:a",
        "libmp3lame",
        "-b:a",
        "320k",
        "-minrate",
        "320k",
        "-maxrate",
        "320k",
        "-bufsize",
        "640k",
        "-write_xing",
        "0",
        str(out),
    ]
    cp = run_checked(cmd)
    if cp.returncode == 0 and out.exists() and out.stat().st_size > 0:
        return ("ok", track, "")

    err = cp.stderr.strip() or "ffmpeg_failed"
    return ("error", track, err)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Deduplicate DJ_YES.xlsx and transcode to MP3 LAME 320 CBR")
    parser.add_argument("--input-xlsx", type=Path, default=Path("/Users/georgeskhawam/Desktop/DJ_YES.xlsx"))
    parser.add_argument("--sheet", type=str, default=None, help="Worksheet name (default: first sheet)")
    parser.add_argument("--output-root", type=Path, default=Path("/Volumes/MUSIC/DJ_YES"))
    parser.add_argument("--manifest-dir", type=Path, default=None, help="Directory for CSV/JSON manifests")
    parser.add_argument("--jobs", type=int, default=max(1, (os.cpu_count() or 4) // 2))
    parser.add_argument("--estimate-sample", type=int, default=160, help="Sample size for duration/size estimate")
    parser.add_argument("--benchmark-sample", type=int, default=14, help="Sample size for encode speed benchmark")
    parser.add_argument("--benchmark-timeout-sec", type=int, default=150)
    parser.add_argument("--seed", type=int, default=20260209)
    parser.add_argument("--estimate-only", action="store_true", help="Only compute estimates and manifests")
    parser.add_argument("--dry-run", action="store_true", help="No transcode, but write manifests and estimates")
    parser.add_argument("--overwrite", action="store_true", help="Overwrite existing output files")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    check_ffmpeg()

    if not args.input_xlsx.exists():
        raise RuntimeError(f"Input XLSX not found: {args.input_xlsx}")

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    manifest_dir = args.manifest_dir or (args.output_root / f"_manifests_{timestamp}")

    tracks, dropped_missing, _headers = load_tracks(args.input_xlsx, args.sheet)
    deduped, dropped_dupes = dedupe_tracks(tracks)
    assign_output_paths(deduped, args.output_root)

    kept_rows = [
        {
            "row_num": t.row_num,
            "source": t.source,
            "external_id": t.external_id,
            "dedupe_key": " | ".join(t.dedupe_key),
            "source_path": str(t.source_path),
            "output_path": str(t.output_path),
        }
        for t in deduped
    ]

    write_csv(
        manifest_dir / "kept_tracks.csv",
        kept_rows,
        ["row_num", "source", "external_id", "dedupe_key", "source_path", "output_path"],
    )
    write_csv(
        manifest_dir / "dropped_missing_on_disk.csv",
        dropped_missing,
        ["row_num", "reason", "path"],
    )
    write_csv(
        manifest_dir / "dropped_duplicates.csv",
        dropped_dupes,
        ["dedupe_key", "kept_row", "kept_path", "dropped_row", "dropped_path", "reason"],
    )

    est = estimate_duration_and_size(deduped, sample_size=args.estimate_sample, sample_seed=args.seed)
    bench = benchmark_realtime_factor(
        deduped,
        sample_size=args.benchmark_sample,
        sample_seed=args.seed,
        timeout_sec=args.benchmark_timeout_sec,
    )
    runtime = estimate_runtime_hours(est["total_duration_sec_est"], bench["avg_realtime_x"], jobs=max(1, args.jobs))

    output_root_existing = args.output_root if args.output_root.exists() else args.output_root.parent
    disk_total, disk_used, disk_free = shutil.disk_usage(output_root_existing)

    summary = {
        "input_xlsx": str(args.input_xlsx),
        "worksheet": args.sheet,
        "output_root": str(args.output_root),
        "jobs": max(1, args.jobs),
        "counts": {
            "input_rows_kept_on_disk": len(tracks),
            "dropped_missing_on_disk": len(dropped_missing),
            "dropped_as_duplicates": len(dropped_dupes),
            "final_deduped_tracks": len(deduped),
        },
        "estimate": est,
        "benchmark": bench,
        "runtime_hours_estimate": runtime,
        "disk": {
            "total_gb": disk_total / (1024**3),
            "used_gb": disk_used / (1024**3),
            "free_gb": disk_free / (1024**3),
            "free_minus_estimated_output_gb": (disk_free - est["output_bytes_est"]) / (1024**3),
        },
    }

    manifest_dir.mkdir(parents=True, exist_ok=True)
    summary_path = manifest_dir / "summary.json"
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")

    print(f"Manifest directory: {manifest_dir}")
    print(json.dumps(summary, indent=2))

    if args.estimate_only or args.dry_run:
        print("Estimate/dry-run mode complete. No transcoding performed.")
        return 0

    total = len(deduped)
    completed = 0
    successes = 0
    skipped_existing = 0
    failures: List[Dict[str, object]] = []

    with ThreadPoolExecutor(max_workers=max(1, args.jobs)) as pool:
        futures = [pool.submit(transcode_one, t, args.overwrite) for t in deduped]
        for future in as_completed(futures):
            status, track, error = future.result()
            completed += 1
            if status == "ok":
                successes += 1
            elif status == "skipped_existing":
                skipped_existing += 1
            else:
                failures.append(
                    {
                        "row_num": track.row_num,
                        "source_path": str(track.source_path),
                        "output_path": str(track.output_path),
                        "error": error,
                    }
                )

            if completed % 100 == 0 or completed == total:
                print(f"Progress: {completed}/{total} (ok={successes}, skipped={skipped_existing}, failed={len(failures)})")

    if failures:
        write_csv(
            manifest_dir / "transcode_failures.csv",
            failures,
            ["row_num", "source_path", "output_path", "error"],
        )

    final = {
        "total": total,
        "ok": successes,
        "skipped_existing": skipped_existing,
        "failed": len(failures),
        "failure_manifest": str(manifest_dir / "transcode_failures.csv") if failures else None,
    }
    print(json.dumps(final, indent=2))
    return 0 if not failures else 2


if __name__ == "__main__":
    raise SystemExit(main())
